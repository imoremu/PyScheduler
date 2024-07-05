import openpyxl
from openpyxl.utils import get_column_letter

import pytest
import pandas as pd
import os
from tempfile import NamedTemporaryFile

from freezegun import freeze_time

from utils.date_utils import safe_to_datetime
from utils.app_config import AppConfig
from utils.util_constants import CONF_DAYFIRST
AppConfig()[CONF_DAYFIRST] = True

from datetime import datetime
from pytest_bdd import scenarios, given, when, then
from scheduler.excel.excel_scheduler import update_excel_schedule

from scheduler.project_scheduler_constants import (
    TASK_ID, TASK_GOAL, TASK_PRIORITY, TASK_ESTIMATION, TASK_RESOURCES_MAX,
    TASK_RESTRICTION, TASK_REMAINING, TASK_START_DATE, TASK_END_DATE,
)

scenarios('../features/update_excel_schedule.feature')

@pytest.fixture
@freeze_time("2024-05-13")
def setup_excel_file():
    pd.set_option('display.max_columns', None)
    pd.set_option('display.expand_frame_repr', False)

    def _create_file(scenario_name):
        TASK_TEAM = "Team"
        TASK_GROUP = "Group"
        TASK_INFO_COLUMN = "Info Column"

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            schedule_sheet = wb.create_sheet(title="Schedule")
            schedule_sheet.append([
                TASK_ID, TASK_GOAL, TASK_PRIORITY, TASK_ESTIMATION, TASK_RESOURCES_MAX,
                TASK_RESTRICTION, TASK_REMAINING, TASK_START_DATE, TASK_END_DATE,
                TASK_TEAM, TASK_GROUP, TASK_INFO_COLUMN
            ])

            if scenario_name == "Plan two tasks with the same priority":
                schedule_sheet.append(["1", "Goal1", 1, 25, 3, "", 10, "", "", "TeamA", "GroupA", "Info1"])
                schedule_sheet.append(["2", "Goal2", 1, 25, 3, "", 10, "", "", "TeamA", "GroupA", "Info2"])
            elif scenario_name == "Plan tasks with different priorities":
                schedule_sheet.append(["1", "Goal1", 1, 25, 3, "", 25, "", "", "TeamA", "GroupA", "Info1"])
                schedule_sheet.append(["2", "Goal2", 2, 25, 3, "", 25, "", "", "TeamA", "GroupA", "Info2"])
            elif scenario_name == "Task blocking another task":
                schedule_sheet.append(["1", "Goal1", 1, 25, 3, "", 25, "", "", "TeamA", "GroupA", "Info1"])
                schedule_sheet.append(["2", "Goal2", 2, 25, 3, "1", 25, "", "", "TeamB", "GroupB", "Info2"])
            elif scenario_name == "Task with a date restriction":
                schedule_sheet.append(["1", "Goal1", 1, 25, 3, "2024-05-25", 25, "", "", "TeamA", "GroupA", "Info1"])
                schedule_sheet.append(["2", "Goal2", 2, 25, 3, "", 25, "", "", "TeamA", "GroupA", "Info2"])
            
            max_row = schedule_sheet.max_row
            max_column = schedule_sheet.max_column            
            last_column_letter = get_column_letter(max_column)
            table_ref = f"A1:{last_column_letter}{max_row}"

            table = openpyxl.worksheet.table.Table(displayName="T_Schedule", ref=table_ref)
            schedule_sheet.add_table(table)

            available_resources_sheet = wb.create_sheet(title="Available Resources")
            available_resources_sheet.append([TASK_GOAL, TASK_TEAM, TASK_GROUP,"12-05-2024","19-05-2024","26-05-2024","02-06-2024"])
            available_resources_sheet.append(["*", "TeamA", "GroupA", 3, 1, 3, 4])
            available_resources_sheet.append(["*", "TeamB", "GroupB", 4, 1, 4, 6])

            max_row = available_resources_sheet.max_row
            max_column = available_resources_sheet.max_column            
            last_column_letter = get_column_letter(max_column)
            table_ref = f"A1:{last_column_letter}{max_row}"

            table = openpyxl.worksheet.table.Table(displayName="T_Available_Resources", ref=table_ref)
            available_resources_sheet.add_table(table)

            used_resources_sheet = wb.create_sheet(title="Used Resources")
            used_resources_sheet.append([TASK_GOAL, TASK_TEAM, TASK_GROUP, TASK_INFO_COLUMN, "12-05-2024","19-05-2024","26-05-2024","02-06-2024"])
            used_resources_sheet.append(["Goal1", "TeamA", "GroupA", "", 2, 2, 2, 2])

            max_row = used_resources_sheet.max_row  
            max_column = used_resources_sheet.max_column
            last_column_letter = get_column_letter(max_column)
            table_ref = f"A1:{last_column_letter}{max_row}"

            table = openpyxl.worksheet.table.Table(displayName="T_Used_Resources", ref=table_ref)
            used_resources_sheet.add_table(table)

            wb.save(tmp.name)
            return tmp.name
    return _create_file

@given('an Excel file with tasks having the same priority', target_fixture="excel_file")
def given_excel_file_same_priority(setup_excel_file):
    return setup_excel_file("Plan two tasks with the same priority")

@given('an Excel file with tasks having different priorities', target_fixture="excel_file")
def given_excel_file_different_priorities(setup_excel_file):
    return setup_excel_file("Plan tasks with different priorities")

@given('an Excel file with tasks where one task blocks the start of another', target_fixture="excel_file")
def given_excel_file_task_blocking(setup_excel_file):
    return setup_excel_file("Task blocking another task")

@given('an Excel file with tasks where one task has a start date restriction', target_fixture="excel_file")
def given_excel_file_date_restriction(setup_excel_file):
    return setup_excel_file("Task with a date restriction")

@when('I update the task schedule in the Excel file with value', target_fixture="when_update_task_schedule")
@freeze_time("2024-05-13")
def when_update_task_schedule(excel_file, tmp_path):
    input_path = excel_file
    # Remove info column from AppConfig
    if "infocolumn" in AppConfig():
        del AppConfig()["infocolumn"]

    output_path = tmp_path / "test_output.xlsx"

    update_excel_schedule(input_path, output_path)

    return output_path

@when('I update the task schedule in the Excel file with formula', target_fixture="when_update_task_schedule")
@freeze_time("2024-05-13")
def when_update_task_schedule_with_formula(excel_file, tmp_path):
    AppConfig()["infocolumn"] = '=INDEX(T_Schedule[{infocolumn}], MATCH([Goal], T_Schedule[Goal], 0),1)'

    input_path = excel_file        
    output_path = tmp_path / "test_output.xlsx"

    update_excel_schedule(input_path, output_path)

    return output_path

@then('the tasks should be planned with the same end date')
def then_verify_same_end_date(when_update_task_schedule):
    output_path = when_update_task_schedule
    assert os.path.exists(output_path)

    wb = openpyxl.load_workbook(output_path)
    schedule_sheet = wb["Schedule"]

    data = pd.DataFrame(schedule_sheet.iter_rows(values_only=True),
                        columns=[cell.value for cell in schedule_sheet[1]])
    
    data = data.iloc[1:]  # Eliminamos la cabecera

    task1_end_date = data.loc[data[TASK_GOAL] == "Goal1"][TASK_END_DATE].values[0]
    task2_end_date = data.loc[data[TASK_GOAL] == "Goal2"][TASK_END_DATE].values[0]

    assert task1_end_date == task2_end_date
    assert safe_to_datetime(task1_end_date) == safe_to_datetime("26-05-2024")

@then('the task with the higher priority should be planned first')
def then_verify_priority_order(when_update_task_schedule):
    output_path = when_update_task_schedule
    assert os.path.exists(output_path)

    wb = openpyxl.load_workbook(output_path)
    schedule_sheet = wb["Schedule"]
    data = pd.DataFrame(schedule_sheet.iter_rows(values_only=True),
                        columns=[cell.value for cell in schedule_sheet[1]])
    data = data.iloc[1:]  

    assert safe_to_datetime(data.iloc[0][TASK_START_DATE]) < safe_to_datetime(data.iloc[1][TASK_START_DATE])
    assert safe_to_datetime(data.loc[data[TASK_GOAL] == "Goal1"][TASK_START_DATE].values[0]) == safe_to_datetime("19-05-2024")
    assert safe_to_datetime(data.loc[data[TASK_GOAL] == "Goal2"][TASK_START_DATE].values[0]) == safe_to_datetime("02-06-2024")

@then('the blocked task should not start until the blocking task is complete')
def then_verify_blocked_task(when_update_task_schedule):
    output_path = when_update_task_schedule
    assert os.path.exists(output_path)

    wb = openpyxl.load_workbook(output_path)
    schedule_sheet = wb["Schedule"]
    data = pd.DataFrame(schedule_sheet.iter_rows(values_only=True),
                        columns=[cell.value for cell in schedule_sheet[1]])
    data = data.iloc[1:] 

    blocking_task_end_date = data.loc[data[TASK_GOAL] == "Goal1"][TASK_END_DATE].values[0]
    blocked_task_start_date = data.loc[data[TASK_GOAL] == "Goal2"][TASK_START_DATE].values[0]
    
    assert safe_to_datetime(blocked_task_start_date) >= safe_to_datetime(blocking_task_end_date)    

@then('the task should not start until the given start date')
def then_verify_date_restriction(when_update_task_schedule):
    output_path = when_update_task_schedule
    assert os.path.exists(output_path)

    wb = openpyxl.load_workbook(output_path)
    
    schedule_sheet = wb["Schedule"]
    data = pd.DataFrame(schedule_sheet.iter_rows(values_only=True),
                        columns=[cell.value for cell in schedule_sheet[1]])
    data = data.iloc[1:] 

    task_start_date = data.iloc[0][TASK_START_DATE]
    assert safe_to_datetime(task_start_date) >= safe_to_datetime("25-05-2024")


@then('the Info Column should be filled with the correct Excel formulas')
def then_verify_info_column_with_formula(when_update_task_schedule):
    output_path = when_update_task_schedule
    assert os.path.exists(output_path)    

    wb = openpyxl.load_workbook(output_path)
    used_resources_sheet = wb["Used Resources"]

    total_data = pd.DataFrame(used_resources_sheet.iter_rows(values_only=True),
                        columns=[cell.value for cell in used_resources_sheet[1]])
    
    data = total_data.iloc[1:]  
        
    info_column_formula = data.loc[data[TASK_GOAL] == "Goal2"]["Info Column"].values[0]
    
    expected_formula = '=INDEX(T_Schedule[Info Column], MATCH([Goal], T_Schedule[Goal], 0),1)'
    assert info_column_formula == expected_formula, f"Expected formula did not match. Found: {info_column_formula}"

@then('the Info Column should be filled with the correct value')
def then_verify_info_column_with_value(when_update_task_schedule):
    output_path = when_update_task_schedule
    assert os.path.exists(output_path)

    wb = openpyxl.load_workbook(output_path)
    used_resources_sheet = wb["Used Resources"]

    total_data = pd.DataFrame(used_resources_sheet.iter_rows(values_only=True),
                        columns=[cell.value for cell in used_resources_sheet[1]])
    
    data = total_data.iloc[1:]  
        
    info_column_formula = data.loc[data[TASK_GOAL] == "Goal2"]["Info Column"].values[0]
    
    expected_formula = 'Info2'
    assert info_column_formula == expected_formula, f"Expected formula did not match. Found: {info_column_formula}"