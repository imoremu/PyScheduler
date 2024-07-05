# Copyright (c) 2024 - Iván Moreno 
#  
# This software is licensed under the MIT License.
# 
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# 
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
# 
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import argparse
import pandas as pd
import openpyxl
from copy import copy

from openpyxl.utils import range_boundaries
from openpyxl.utils import get_column_letter

from scheduler.project_scheduler_constants import TASK_END_DATE, TASK_GOAL, TASK_START_DATE, USED_RESOURCE_GOAL
from scheduler.project_resource_manager import ProjectResourceManager
from scheduler.project_task_scheduler import TaskManager

from utils.app_config import AppConfig
from utils.date_utils import safe_to_datetime

from utils.logger import create_logger
from openpyxl.formatting.formatting import ConditionalFormattingList

logger = create_logger(__name__)

def load_table(sheet, table_name, valueorformula='value'):
    """ 
    Load the data from an Excel table into a pandas DataFrame.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet containing the table to be loaded.
        table_name (str): The name of the table to be loaded.
    """
    table = sheet.tables[table_name]
    data = sheet[table.ref]
    headers = [cell.value for cell in data[0]]

    if valueorformula == 'value':
        data_values = [[cell.value for cell in row] for row in data[1:]]
    elif valueorformula == 'formula':
        data_values = [[cell.formula for cell in row] for row in data[1:]]
    else:
        raise ValueError("valueorformula must be 'value' or 'formula'") 
    
    return pd.DataFrame(data_values, columns=headers)

def apply_conditional_formatting(sheet, start_row, end_row):
    """
    Apply existing conditional formatting rules to new rows.
    Assumes that rules apply to entire columns.
    """        
    rules = sheet.conditional_formatting._cf_rules    
    new_rules = []
    
    for range, formatting_rules in rules.items():   
        for rng in range.sqref:
            min_col, min_row, max_col, max_row = range_boundaries(str(rng))
            
            new_range = f"{sheet.cell(row=start_row, column=min_col).coordinate}:{sheet.cell(row=end_row, column=max_col).coordinate}"
            
            for rule in formatting_rules:
                new_rule = copy(rule)

                new_rule.ranges = [new_range]
            
                new_rules.append(new_rule)

    # Clear existing rules in target range to avoid duplication or overlap (rules is a OrderedDict, so order is preserved)
    sheet.conditional_formatting = ConditionalFormattingList()
    
    for rule in new_rules:
        sheet.conditional_formatting.add(rule.ranges[0], rule)

def copy_row_styles(sheet, source_row, target_row, min_col, max_col):
    """
    Copy the style of all cells in a source row to a target row within specified columns.
    """
    for col in range(min_col, max_col + 1):
        source_cell = sheet.cell(row=source_row, column=col)
        target_cell = sheet.cell(row=target_row, column=col)
        # Copiar estilo
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment) 

def update_table(sheet, table_name, data, columns=None):
    """
    Update the data in an Excel table.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet containing the table to be updated.
        table_name (str): The name of the table to be updated.
        data (pandas.DataFrame): The data to be updated in the table.
        columns (list of str, optional): List of column names to update. Updates all columns if None.
    """
    table = sheet.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    
    if columns is not None:
        update_cols = [data.columns.get_loc(col) + min_col for col in columns if col in data.columns]
    else:
        update_cols = range(min_col, max_col + 1)

    new_max_row = min_row + data.shape[0]
    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"

    if new_max_row > max_row:
        for i in range(max_row + 1, new_max_row + 1):
            sheet.append([None] * (max_col - min_col + 1))
            copy_row_styles(sheet, max_row, i, min_col, max_col)

    for r_idx, row in enumerate(sheet.iter_rows(min_row=min_row + 1, max_row=new_max_row, min_col=min_col, max_col=max_col), start=1):
        for c_idx, cell in enumerate(row, start=0):
            if cell.column in update_cols:
                col_name = data.columns[c_idx]
                cell_value = data.iloc[r_idx - 1, data.columns.get_loc(col_name)]
                cell.value = cell_value

    if new_max_row < max_row:
        for row in sheet.iter_rows(min_row=new_max_row + 1, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.column in update_cols:
                    cell.value = None

    apply_conditional_formatting(sheet, min_row, new_max_row)


def update_excel_schedule(file_path, output_path):
    """
    Update the schedule in an Excel file.

    Args:
        file_path (str): Path to the Excel file containing the schedule to be updated.
        output_path (str): Path to the output Excel file.
    
    """
    try:
        config = AppConfig()

        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)

        # Load the sheets
        tasks_sheet = wb['Schedule']
        tasks_sheet_final = wb_formulas['Schedule']

        available_resources_sheet = wb['Available Resources']
        
        used_resources_sheet = wb_formulas['Used Resources']

        # Convert the Excel tables to DataFrames
        tasks_df = load_table(tasks_sheet, 'T_Schedule')
        available_resources_df = load_table(available_resources_sheet, 'T_Available_Resources')
        used_resources_df = load_table(used_resources_sheet, 'T_Used_Resources')

        # Initialize the ResourceManager and TaskManager
        resource_manager = ProjectResourceManager(available_resources_df, used_resources_df, tasks_df)
        task_manager = TaskManager(tasks_df, resource_manager)

        # Define the dates you want to update the schedule for
        dates = [
            col for col in available_resources_df.columns if safe_to_datetime(col, errors='coerce') is not pd.NaT
        ]

        # Update the task schedule
        updated_tasks_df = task_manager.update_task_schedule(dates)

        # Update the 'T_Schedule' with updated tasks DataFrame
        update_table(tasks_sheet_final, 'T_Schedule', updated_tasks_df, columns=[TASK_START_DATE, TASK_END_DATE])
        update_table(used_resources_sheet, 'T_Used_Resources', task_manager.resource_manager.used_resources_manager.used_resources_df)

        # Save the workbook
        wb_formulas.save(output_path)
        logger.info(f"Successfully updated {output_path}")
    
    except Exception as e:

        logger.error(f"An error occurred while updating the schedule: {e}")
        raise   

if __name__ == "__main__":
    """
    Update the schedule in an Excel file.

    Usage:
    python excel_scheduler.py -i <input_file> -o <output_file> [-p <period>] [-v <holidays>] [-d] [-c <conffile>]

    Arguments:
    -i, --input: Path to the Excel file containing the schedule to be updated.
    -o, --output: Path to the output Excel file. Defaults to the same as the input if not specified.
    -p, --period: Number of useful resource days between dates. Possible values: integer value or 'auto'. Default: 5.
    -v, --holidays: List of bank holidays in JSON format or path to a JSON file with the list. Default: Empty list.
    -d, --dayfirst: Set if the dates in the Excel file are in 'DD/MM/YYYY' format. Default is True.
    -c, --conffile: Provide a file with arguments coming from a json file as a dictionary.

    
    Example:
    From attributes:
    
        python excel_scheduler.py -i schedule.xlsx -o updated_schedule.xlsx -p 5 -v holidays.json -d False

    From a configuration file:

        python excel_scheduler.py -c config.json


    """
    parser = argparse.ArgumentParser(description="Update the schedule in an Excel file.")
    parser.add_argument('-i', '--input', help="Path to the Excel file containing the schedule to be updated.")
    parser.add_argument('-o', '--output', help="Path to the output Excel file. Defaults to the same as the input if not specified.")   
    parser.add_argument('-p', '--period', default=5, help="Number of useful resource days between dates. Possible values: integer value or 'auto'. Default: 5.") 
    parser.add_argument('-v', '--holidays', default=[], help="List of bank holidays in JSON format or path to a JSON file with the list. Default: Empty list.")     
    parser.add_argument('-d', '--dayfirst', default=True, action='store_true', help="Set if the dates in the Excel file are in 'DD/MM/YYYY' format. Default is True.")
    parser.add_argument('-c', '--conffile', default="", action='store_true', help="Provide a file with arguments coming from a json file as a dictionary.")
    parser.add_argument('-l', '--log', default="INFO", help="Set the logging level. Default is INFO.")
    parser.add_argument('-ic', '--infocolumn', default=f"=INDEX(T_Schedule[{{infocolumn}}], MATCH([{TASK_GOAL}], T_Schedule[{TASK_GOAL}], 0),1)", help="Set the info columns values. Default is INDEX(T_Schedule[{attr}], MATCH([{TASK_GOAL}], T_Schedule[{TASK_GOAL}], 0),1).")

    args = parser.parse_args()
    config = AppConfig()

    config.load_args(args)

    logger.info(f"Starting the update process for {config["input"]}")
    update_excel_schedule(config["input"], config["output"])
    logger.info(f"Finished updating {config["input"]}, output saved to {config["output"]}")